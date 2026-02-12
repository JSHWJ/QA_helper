from io import BytesIO

import pandas as pd

from pathlib import Path
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from app_modules.storage_utils import current_timestamp_text, export_dir_path, reserve_next_version


# 파일 내보내기 기능 모듈


def _column_index_map(dataframe: pd.DataFrame) -> dict[str, int]:
    return {str(name): idx + 1 for idx, name in enumerate(list(dataframe.columns))}


def _col_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


def _norm_formula(cell_ref: str) -> str:
    base = (
        f'TRIM(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({cell_ref},CHAR(13)&CHAR(10),CHAR(10)),'
        'CHAR(13),CHAR(10)),CHAR(160)," "))'
    )
    return (
        f'=IF(AND(LEN({base})>=2,LEFT({base},1)=RIGHT({base},1),'
        f'OR(CODE(LEFT({base},1))=34,CODE(LEFT({base},1))=39)),'
        f'TRIM(MID({base},2,LEN({base})-2)),{base})'
    )


def _apply_excel_match_formulas_and_styles(worksheet, dataframe: pd.DataFrame):
    if dataframe.empty:
        return

    idx = _column_index_map(dataframe)
    max_row = len(dataframe) + 1
    if max_row < 2:
        return

    required_columns = [
        "Dictionary English",
        "Dictionary Korean",
        "Dictionary Russian",
        "en.json",
        "ko.json",
        "ru.json",
        "KO_Match",
        "EN_Match",
        "RU_Match",
        "Overall_Match",
    ]
    if any(col not in idx for col in required_columns):
        return

    helper_specs = [
        ("Dictionary Korean", "__NORM_DICTIONARY_KOREAN"),
        ("ko.json", "__NORM_KO_JSON"),
        ("Dictionary English", "__NORM_DICTIONARY_ENGLISH"),
        ("en.json", "__NORM_EN_JSON"),
        ("Dictionary Russian", "__NORM_DICTIONARY_RUSSIAN"),
        ("ru.json", "__NORM_RU_JSON"),
    ]

    helper_start_idx = worksheet.max_column + 1
    helper_idx = {}
    for offset, (_, helper_name) in enumerate(helper_specs):
        col_idx = helper_start_idx + offset
        helper_idx[helper_name] = col_idx
        worksheet.cell(row=1, column=col_idx, value=helper_name)

    for row in range(2, max_row + 1):
        for source_col_name, helper_name in helper_specs:
            source_cell = f"{_col_letter(idx[source_col_name])}{row}"
            helper_cell = worksheet.cell(row=row, column=helper_idx[helper_name])
            helper_cell.value = _norm_formula(source_cell)

        ko_dict = f"{_col_letter(helper_idx['__NORM_DICTIONARY_KOREAN'])}{row}"
        ko_json = f"{_col_letter(helper_idx['__NORM_KO_JSON'])}{row}"
        en_dict = f"{_col_letter(helper_idx['__NORM_DICTIONARY_ENGLISH'])}{row}"
        en_json = f"{_col_letter(helper_idx['__NORM_EN_JSON'])}{row}"
        ru_dict = f"{_col_letter(helper_idx['__NORM_DICTIONARY_RUSSIAN'])}{row}"
        ru_json = f"{_col_letter(helper_idx['__NORM_RU_JSON'])}{row}"

        worksheet.cell(row=row, column=idx["KO_Match"]).value = (
            f'=IF({ko_json}="","파일없음",'
            f'IF({ko_dict}="","N",'
            f'IF(ISNUMBER(SEARCH(","&{ko_json}&",",","&SUBSTITUTE({ko_dict},", ",",")&",")),"Y","N")))'
        )
        worksheet.cell(row=row, column=idx["EN_Match"]).value = (
            f'=IF({en_json}="","파일없음",'
            f'IF({en_dict}="","N",'
            f'IF(ISNUMBER(SEARCH(","&{en_json}&",",","&SUBSTITUTE({en_dict},", ",",")&",")),"Y","N")))'
        )
        worksheet.cell(row=row, column=idx["RU_Match"]).value = (
            f'=IF({ru_json}="","파일없음",'
            f'IF({ru_dict}="","N",'
            f'IF(ISNUMBER(SEARCH(","&{ru_json}&",",","&SUBSTITUTE({ru_dict},", ",",")&",")),"Y","N")))'
        )

        ko_match_cell = f"{_col_letter(idx['KO_Match'])}{row}"
        en_match_cell = f"{_col_letter(idx['EN_Match'])}{row}"
        ru_match_cell = f"{_col_letter(idx['RU_Match'])}{row}"
        worksheet.cell(row=row, column=idx["Overall_Match"]).value = (
            f'=IF(OR({ko_match_cell}="파일없음",{en_match_cell}="파일없음",{ru_match_cell}="파일없음"),"파일없음",'
            f'IF(AND({ko_match_cell}="Y",{en_match_cell}="Y",{ru_match_cell}="Y"),"Y","N"))'
        )

    yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    gray_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    for match_col in ["KO_Match", "EN_Match", "RU_Match", "Overall_Match"]:
        letter = _col_letter(idx[match_col])
        cell_range = f"{letter}2:{letter}{max_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"N"'], fill=yellow_fill),
        )
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="equal", formula=['"파일없음"'], fill=gray_fill),
        )

    for helper_col_idx in helper_idx.values():
        worksheet.column_dimensions[_col_letter(helper_col_idx)].hidden = True

def dataframe_to_excel_bytes(dataframe: pd.DataFrame) -> bytes:
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Compare")
        worksheet = writer.book["Compare"]
        _apply_excel_match_formulas_and_styles(worksheet, dataframe)
    return bio.getvalue()


def save_dataframe_to_export_folder(dataframe: pd.DataFrame, base_name: str, export_dir: str | None = None):
    version = reserve_next_version()
    ts = current_timestamp_text()
    target_dir = Path(export_dir) if export_dir else export_dir_path()
    target_dir.mkdir(parents=True, exist_ok=True)

    csv_path = target_dir / f"{base_name}_v{version}_{ts}.csv"
    xlsx_path = target_dir / f"{base_name}_v{version}_{ts}.xlsx"

    dataframe.to_csv(csv_path, index=False, encoding="utf-8-sig")
    excel_bytes = dataframe_to_excel_bytes(dataframe)
    xlsx_path.write_bytes(excel_bytes)
    return csv_path, xlsx_path, version, ts
